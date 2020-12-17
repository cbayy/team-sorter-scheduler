from openpyxl import *
import time
import operator
from tkinter import * 
from tkinter import filedialog 

class Student:
    team = None
    def __init__(self, name, gender, group, score):
        self.name = name
        self.gender = gender
        self.group = group
        self.score = score

def main():
    filename = browseFiles()
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    array = getCells(sheet, workbook)

    groupNum = sendMessage(array)
    genderedSort = determineGender()
    sortedArray = runSorter(genderedSort, array, groupNum)
    teamScore = calculateScores(sortedArray, groupNum)

    writeTotalData(teamScore, sheet, workbook)
    writeData(array, sheet, workbook, filename)
    workbook.save(filename)
    endProcess(filename)

def browseFiles(): 
    filename = filedialog.askopenfilename(initialdir = "/Desktop/", 
                                          title = "Select a File", 
                                          filetypes = [("Excel file","*.xlsx"),("Excel file", "*.xls")])
    label_file_explorer.configure(text="File Opened: "+filename) 
    window.destroy()
    return filename

def getCells(sheet, workbook):
    array = []
    clearCells(100, sheet)
    for value in sheet.iter_rows(
        min_row=2,
        max_col=4,
        values_only=True,
    ):
        if(value[0] != None):
            temp = Student(value[0],value[1],value[2], value[3])
            array.append(temp)
        else:
            break
    return array
    
def clearCells(array, sheet):
    for s in range(array):
        sheet.cell(row=s+2, column=7).value = None
        sheet.cell(row=s+2, column =6).value = None
        sheet.cell(row=s+2, column =8).value = None

    for g in range(array):
        sheet.cell(row=g+2, column = 10).value = None
        sheet.cell(row=g+2, column = 11).value = None

def sendMessage(array):
    print("The number of student in this class: " + str(len(array)))
    print("Suitable # of groups: ")
    for n in range(int(len(array))):
        if(n > 1):
            if((int(len(array)) % n) == 0):
                print("\t\t\t" + str(n))
    groupNum = input("Enter number of groups: ")
    print("Confirmed")
    return groupNum

def determineGender():
    genderedSort = input("Sort by gender [Y/N] (warning: teams are less likely to be even by score): ")
    return genderedSort

def runSorter(genderedSort, array, groupNum):
    if(genderedSort == "N"):
        array = sortStudents(array, groupNum)
    else:
        array = sortStudentsGender(array, groupNum)
    return array

def sortStudents(array, groupNum):
    array = sorted(array, key=lambda student: student.score)
    g = 0
    ascending = TRUE
    for s in range(len(array)):
        if(ascending == TRUE):
            if(g == int(groupNum)):
                ascending = FALSE
                array[s].team = g
                g = g - 1
            else:
                array[s].team = g
                g = g + 1
        if(ascending == FALSE):
            if(g == 0):
                ascending = TRUE
                array[s].team = g
            else:
                array[s].team = g
                g = g - 1
    return array

def sortStudentsGender(array, groupNum):
    array.sorted(array, key=lambda student: student.gender)
    for s in range(len(array)):
        array[s].team = s
    return array

def calculateScores(array, groupNum):
    teamScore = []
    for x in range(int(groupNum)):
        teamScore.append(0)

    for s in range(len(array)):
        for g in range(len(teamScore)):
            if(g == array[s].team):
                teamScore[g] = teamScore[g] + array[s].score
    return teamScore
         
def writeData(array, sheet, workbook, filename):
    sheet.cell(row=1,column=7).value = "Name"
    sheet.cell(row=1,column=8).value = "Team"
    
    array.sort(key = lambda student: student.team)
    
    n = 0
    for s in range(len(array)):

        if(array[s].team != n):
            sheet.cell(row=s+2+n, column=7).value = ""
            sheet.cell(row=s+2+n, column=8).value = ""
            n = n + 1
        sheet.cell(row=s+2+n, column =7).value = array[s].name
        sheet.cell(row=s+2+n, column =8).value = array[s].team + 1
            
def writeTotalData(teamScore, sheet, workbook):
    sheet.cell(row=1,column=10).value = "Team number"
    sheet.cell(row=1,column=11).value = "Total score"

    for g in range(len(teamScore)):
        sheet.cell(row=g+2, column = 10).value = g + 1
        sheet.cell(row=g+2, column = 11).value = teamScore[g] 

def endProcess(filename):
    print("\nProcessing finished. Additional data added to: "  + filename)
    for i in range(9, 0, -1):
        print("Closing window in: " + str(i), end = '\r')
        time.sleep(1)
    quit()

window = Tk() 
window.title('File Explorer') 
window.geometry("700x200") 
window.config(background = "white") 

label_file_explorer = Label(window,  
                            text = "Please select excel file", 
                            width = 100, height = 4,  
                            fg = "black")    
   
button_explore = Button(window,  
                        text = "Browse Files", 
                        command = main)  
   
button_exit = Button(window,  
                     text = "Exit", 
                     command = exit)  
   
label_file_explorer.grid(column = 1, row = 1) 
   
button_explore.grid(column = 1, row = 2) 
   
button_exit.grid(column = 1,row = 3) 
   
window.mainloop() 